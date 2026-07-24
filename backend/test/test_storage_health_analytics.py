# -*- coding: utf-8 -*-
import csv
import importlib
import importlib.util
import io
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest.mock import ANY, Mock, patch
from zoneinfo import ZoneInfo

import pytest
import sqlalchemy as sa
from alembic.migration import MigrationContext
from alembic.operations import Operations
from openpyxl import load_workbook

import models
import questdb.models  # noqa: F401
from appConfig import Config
from crud import storageAlertCrud, storageHealthAnalyticsCrud
from questdb.database import QuestDBBase
from routers import storage_cluster
from utils.isilonClient import IsilonClient
from utils.netAppClient import NetAppClient


BACKEND_ROOT = Path(__file__).resolve().parents[1]
START = datetime(2026, 7, 1, tzinfo=timezone.utc)
END = datetime(2026, 7, 2, tzinfo=timezone.utc)
SYSTEM_TIMEZONE = ZoneInfo("Asia/Shanghai")


def _analytics():
    return importlib.import_module("services.storageHealthAnalyticsService")


def _query_params(**overrides):
    params = {"start_time": START.isoformat(), "end_time": END.isoformat()}
    params.update(overrides)
    return params


def test_capacity_change_uses_chronological_first_and_last_samples():
    result = _analytics().summarize_capacity(
        [
            {"updated_at": END, "used": 175.0},
            {"updated_at": START, "used": 100.0},
        ]
    )

    assert result == {
        "start_used": 100.0,
        "end_used": 175.0,
        "change": 75.0,
        "change_percent": 75.0,
        "points": [
            {"updated_at": START, "used": 100.0},
            {"updated_at": END, "used": 175.0},
        ],
    }


@pytest.mark.parametrize(
    ("points", "expected"),
    [
        (
            [{"updated_at": START, "used": 0.0}, {"updated_at": END, "used": 5.0}],
            {
                "start_used": 0.0,
                "end_used": 5.0,
                "change": 5.0,
                "change_percent": None,
                "points": [
                    {"updated_at": START, "used": 0.0},
                    {"updated_at": END, "used": 5.0},
                ],
            },
        ),
        (
            [],
            {
                "start_used": None,
                "end_used": None,
                "change": None,
                "change_percent": None,
                "points": [],
            },
        ),
    ],
    ids=["zero-baseline", "empty"],
)
def test_capacity_change_handles_zero_baseline_and_empty_data(points, expected):
    assert _analytics().summarize_capacity(points) == expected


def test_capacity_summary_uses_real_boundaries_instead_of_bucket_maxima():
    points = [
        {"updated_at": START, "used": 150.0},
        {"updated_at": END, "used": 200.0},
    ]

    result = _analytics().summarize_capacity(
        points,
        start_used=100.0,
        end_used=175.0,
    )

    assert result["start_used"] == 100.0
    assert result["end_used"] == 175.0
    assert result["change"] == 75.0
    assert result["change_percent"] == 75.0
    assert result["points"] == points


def test_severity_summary_normalizes_and_merges_alert_sources():
    result = _analytics().summarize_severities(
        [
            {"source": "diskpulse", "severity": "high"},
            {"source": "diskpulse", "severity": "medium"},
            {"source": "netapp", "severity": "emergency"},
            {"source": "netapp", "severity": "error"},
            {"source": "isilon", "severity": "warning"},
            {"source": "isilon", "severity": "informational"},
        ]
    )

    assert result == {
        "counts": {"critical": 2, "error": 1, "warning": 2, "info": 1},
        "total": 6,
        "sources": {
            "diskpulse": {"critical": 1, "error": 0, "warning": 1, "info": 0},
            "netapp": {"critical": 1, "error": 1, "warning": 0, "info": 0},
            "isilon": {"critical": 0, "error": 0, "warning": 1, "info": 1},
        },
    }


def test_alert_crud_preserves_aware_utc_bounds_for_postgresql_queries():
    assert storageHealthAnalyticsCrud._naive(
        datetime(2026, 7, 15, 2, 0, tzinfo=timezone.utc)
    ) == datetime(2026, 7, 15, 2, 0, tzinfo=timezone.utc)
    with pytest.raises(ValueError, match="timezone-aware"):
        storageHealthAnalyticsCrud._naive(datetime(2026, 7, 15, 10, 0))


def test_error_severity_window_matches_the_vendor_system_event_scope(db_session):
    db_session.add(
        models.StorageCluster(
            id=9,
            name="cluster-nine",
            storage_type="netapp",
            storage_host="storage.local",
            is_active=True,
        )
    )
    db_session.add_all(
        [
            models.StorageAlerts(
                storage_cluster_id=9,
                source="netapp",
                external_event_id="vendor-1",
                severity="error",
                alert_level="error",
                updated_at=datetime(2026, 7, 15, 2, 15, tzinfo=timezone.utc),
            ),
            models.StorageAlerts(
                storage_cluster_id=9,
                source="diskpulse",
                external_event_id="diskpulse-1",
                severity="warning",
                alert_level="medium",
                updated_at=datetime(2026, 7, 15, 2, 30, tzinfo=timezone.utc),
            ),
        ]
    )
    db_session.commit()

    rows = storageHealthAnalyticsCrud.get_alert_severities(
        db_session,
        9,
        datetime(2026, 7, 15, 2, 0, tzinfo=timezone.utc),
        datetime(2026, 7, 15, 3, 0, tzinfo=timezone.utc),
    )

    assert {(row["source"], row["severity"]) for row in rows} == {("netapp", "error")}


def test_error_severity_crud_allows_only_vendor_system_event_sources(db_session):
    db_session.add(
        models.StorageCluster(
            id=10,
            name="cluster-ten",
            storage_type="isilon",
            storage_host="storage.local",
            is_active=True,
        )
    )
    for index, source in enumerate(("diskpulse", "netapp", "isilon", "custom"), start=1):
        db_session.add(
            models.StorageAlerts(
                storage_cluster_id=10,
                source=source,
                external_event_id=f"severity-{index}",
                severity="error",
                alert_level="error",
                updated_at=datetime(2026, 7, 15, 2, index, tzinfo=timezone.utc),
            )
        )
    db_session.commit()

    rows = storageHealthAnalyticsCrud.get_alert_severities(
        db_session,
        10,
        datetime(2026, 7, 15, 2, 0, tzinfo=timezone.utc),
        datetime(2026, 7, 15, 3, 0, tzinfo=timezone.utc),
    )

    assert {row["source"] for row in rows} == {"netapp", "isilon"}


def test_general_alert_query_excludes_vendor_system_events(db_session):
    for index, source in enumerate(("diskpulse", "netapp", "isilon"), start=1):
        db_session.add(
            models.StorageAlerts(
                source=source,
                external_event_id=f"event-{index}",
                severity="warning",
                alert_level="warning",
                alert_type="vendor_event" if source != "diskpulse" else "alert",
                description=source,
                updated_at=datetime(2026, 7, 15, 10, index, tzinfo=timezone.utc),
            )
        )
    db_session.commit()

    rows, total = storageAlertCrud.get_storage_alerts(db_session, page=1, size=20)

    assert total == 1
    assert [row.source for row in rows] == ["diskpulse"]


def test_system_events_return_only_vendor_rows_for_the_selected_cluster(db_session):
    db_session.add_all(
        [
            models.StorageCluster(id=7, name="cluster-seven", storage_type="netapp"),
            models.StorageCluster(id=8, name="cluster-eight", storage_type="isilon"),
        ]
    )
    for index, (cluster_id, source) in enumerate(
        ((7, "netapp"), (7, "diskpulse"), (8, "isilon")),
        start=1,
    ):
        db_session.add(
            models.StorageAlerts(
                storage_cluster_id=cluster_id,
                source=source,
                external_event_id=f"system-event-{index}",
                severity="error",
                alert_level="error",
                alert_type="vendor_event",
                description=f"{source}-message",
                related_info={
                    "event_code": f"code-{index}",
                    "object_id": f"object-{index}",
                    "raw": {"node": {"name": "node-a"}},
                },
                updated_at=datetime(2026, 7, 15, 10, index, tzinfo=timezone.utc),
            )
        )
    db_session.commit()

    total, rows = storageHealthAnalyticsCrud.get_system_event_rows(
        db_session,
        7,
            datetime(2026, 7, 15, 10, 0, tzinfo=timezone.utc),
            datetime(2026, 7, 15, 11, 0, tzinfo=timezone.utc),
    )

    assert total == 1
    assert rows == [
        {
            "source": "netapp",
            "severity": "error",
            "event_code": "code-1",
            "object_id": "object-1",
            "object_name": "node-a",
            "object_type": "node",
            "description": "netapp-message",
                "occurred_at": datetime(2026, 7, 15, 10, 1, tzinfo=timezone.utc),
        }
    ]


def test_system_events_filter_before_database_pagination(db_session):
    db_session.add(
        models.StorageCluster(id=11, name="cluster-eleven", storage_type="isilon")
    )
    db_session.add_all(
        [
            models.StorageAlerts(
                storage_cluster_id=11,
                source="isilon",
                external_event_id=f"quota-{index}",
                fingerprint=f"isilon:QUOTA_{index}:node:6",
                severity="warning" if index < 4 else "error",
                alert_level="warning" if index < 4 else "error",
                alert_type="vendor_event",
                description=f"Quota threshold event {index}",
                related_type="node",
                related_info={"event_code": f"QUOTA_{index}", "object_id": "6"},
                updated_at=datetime(2026, 7, 15, 10, index, tzinfo=timezone.utc),
            )
            for index in range(1, 5)
        ]
    )
    db_session.commit()

    total, rows = storageHealthAnalyticsCrud.get_system_event_rows(
        db_session,
        11,
            datetime(2026, 7, 15, 10, 0, tzinfo=timezone.utc),
            datetime(2026, 7, 15, 11, 0, tzinfo=timezone.utc),
        keyword="quota",
        severity="warning",
        page=2,
        page_size=2,
    )

    assert total == 3
    assert [row["event_code"] for row in rows] == ["QUOTA_1"]
    assert rows[0]["object_name"] == "节点 6"
    assert rows[0]["object_id"] == "6"


def test_system_event_service_returns_pagination_metadata():
    analytics = _analytics()
    db = Mock()
    rows = [{"event_code": "QUOTA_1"}]
    with patch.object(
        analytics.storageHealthAnalyticsCrud,
        "get_system_event_rows",
        return_value=(41, rows),
    ) as get_rows:
        result = analytics.get_system_events(
            db,
            11,
            START,
            END,
            keyword="quota",
            severity="warning",
            page=2,
            page_size=20,
        )

    get_rows.assert_called_once_with(
        db,
        11,
        START,
        END,
        keyword="quota",
        severity="warning",
        include_identity=True,
        page=2,
        page_size=20,
    )
    assert result == {"data": rows, "total": 41, "page": 2, "page_size": 20}


def test_top_latency_ranks_by_p95_and_applies_limit():
    rows = [
        {
            "object_id": str(index),
            "object_name": f"volume-{index}",
            "object_type": "volume",
            "p95_latency": float(index),
            "avg_latency": float(index) / 2,
            "max_latency": float(index) + 1,
            "sample_count": index + 1,
        }
        for index in range(12)
    ]

    result = _analytics().rank_top_latency(rows, limit=10)

    assert [row["object_id"] for row in result] == [str(index) for index in range(11, 1, -1)]
    assert all({"p95_latency", "avg_latency", "max_latency", "sample_count"} <= row.keys() for row in result)


def test_top_latency_crud_returns_standard_performance_metrics():
    connection = Mock()
    connection.execute.return_value.all.return_value = [
        ("volume-1", "vol-a", "volume", 9.5, 5.0, 12.0, 3.0, 7.0, 125.0, 4096.0, 8)
    ]
    session = Mock()
    session.__enter__ = Mock(return_value=connection)
    session.__exit__ = Mock(return_value=False)

    with (
        patch.object(storageHealthAnalyticsCrud, "get_storage_config", return_value=Mock()),
        patch.object(storageHealthAnalyticsCrud, "QuestDBSession", return_value=session),
    ):
        rows = storageHealthAnalyticsCrud.get_top_latency_rows(
            Mock(), 7, START, END, 10, "volume", {"volume-1"}
        )

    statement = str(connection.execute.call_args.args[0])
    assert "avg(latency_read) AS avg_read_latency" in statement
    assert "avg(latency_write) AS avg_write_latency" in statement
    assert "avg(iops_total) AS avg_iops" in statement
    assert "avg(throughput_total) AS avg_throughput" in statement
    assert "object_id IN (:object_id_0)" in statement
    assert connection.execute.call_args.args[1]["object_id_0"] == "volume-1"
    assert rows == [
        {
            "object_id": "volume-1",
            "object_name": "vol-a",
            "object_type": "volume",
            "p95_latency": 9.5,
            "avg_latency": 5.0,
            "max_latency": 12.0,
            "avg_read_latency": 3.0,
            "avg_write_latency": 7.0,
            "avg_iops": 125.0,
            "avg_throughput": 4096.0,
            "sample_count": 8,
        }
    ]


def test_hourly_asset_performance_uses_utc_window_and_bound_resource_identity():
    connection = Mock()
    connection.execute.return_value.all.return_value = [
        (START, 6, 1.5, 2.5, 2.0, 240.0, 4096.0),
    ]
    session = Mock()
    session.__enter__ = Mock(return_value=connection)
    session.__exit__ = Mock(return_value=False)

    with (
        patch.object(storageHealthAnalyticsCrud, "get_storage_config", return_value=Mock()),
        patch.object(storageHealthAnalyticsCrud, "QuestDBSession", return_value=session),
    ):
        rows = storageHealthAnalyticsCrud.get_hourly_asset_performance(
            Mock(),
            storage_cluster_id=7,
            asset_type="volume",
            asset_id="volume-9",
            start_time=START,
            end_time=END,
        )

    statement = str(connection.execute.call_args.args[0])
    parameters = connection.execute.call_args.args[1]
    assert "storage_performance_metrics" in statement
    assert "object_type = :asset_type" in statement
    assert "object_id = :asset_id" in statement
    assert "SAMPLE BY 1h" in statement
    assert parameters == {
        "storage_cluster_id": "7",
        "asset_type": "volume",
        "asset_id": "volume-9",
        "start_time": "2026-07-01T00:00:00Z",
        "end_time": "2026-07-02T00:00:00Z",
    }
    assert rows == [{
        "hour_start": START,
        "sample_count": 6,
        "latency_read": 1.5,
        "latency_write": 2.5,
        "latency_total": 2.0,
        "iops_total": 240.0,
        "throughput_total": 4096.0,
    }]


def test_repeated_faults_group_vendor_events_only():
    result = _analytics().group_repeated_faults(
        [
            {"source": "netapp", "fingerprint": "disk.offline:disk-1", "occurred_at": START},
            {"source": "netapp", "fingerprint": "disk.offline:disk-1", "occurred_at": END},
            {"source": "isilon", "fingerprint": "node.down:node-1", "occurred_at": START},
            {"source": "diskpulse", "fingerprint": "quota:group-1", "occurred_at": START},
            {"source": "diskpulse", "fingerprint": "quota:group-1", "occurred_at": END},
        ]
    )

    assert result == [
        {
            "source": "netapp",
            "fingerprint": "disk.offline:disk-1",
            "count": 2,
            "first_occurred_at": START,
            "last_occurred_at": END,
        }
    ]


def test_repeated_faults_group_excludes_unknown_sources():
    result = _analytics().group_repeated_faults(
        [
            {"source": "custom", "fingerprint": "disk.offline:disk-1", "occurred_at": START},
            {"source": "custom", "fingerprint": "disk.offline:disk-1", "occurred_at": END},
            {"source": "isilon", "fingerprint": "node.down:node-1", "occurred_at": START},
            {"source": "isilon", "fingerprint": "node.down:node-1", "occurred_at": END},
        ]
    )

    assert {row["source"] for row in result} == {"isilon"}


def test_repeated_fault_crud_allows_only_netapp_and_isilon(db_session):
    db_session.add(
        models.StorageCluster(
            id=7,
            name="cluster-seven",
            storage_type="netapp",
            storage_host="storage.local",
            is_active=True,
        )
    )
    for source in ("netapp", "custom"):
        for index, occurred_at in enumerate((START, END), start=1):
            db_session.add(
                models.StorageAlerts(
                    storage_cluster_id=7,
                    source=source,
                    external_event_id=f"{source}-{index}",
                    fingerprint=f"{source}:disk.offline:node:node-1",
                    severity="error",
                    alert_level="error",
                    updated_at=occurred_at,
                )
            )
    db_session.commit()

    rows = storageHealthAnalyticsCrud.get_repeated_fault_rows(
        db_session,
        7,
        START,
        END,
    )

    assert {row["source"] for row in rows} == {"netapp"}


def test_top_latency_empty_range_distinguishes_supported_from_never_collected():
    analytics = _analytics()
    db = Mock()
    with (
        patch.object(
            analytics.storageHealthAnalyticsCrud,
            "get_top_latency_rows",
            return_value=[],
        ),
        patch.object(
            analytics.storageHealthAnalyticsCrud,
            "has_performance_metrics",
            side_effect=[True, False],
            create=True,
        ) as has_performance_metrics,
    ):
        supported = analytics.get_top_latency(db, 7, START, END)
        unsupported = analytics.get_top_latency(db, 8, START, END)

    assert [item.args for item in has_performance_metrics.call_args_list] == [
        (db, 7),
        (db, 8),
    ]
    assert supported == {"supported": True, "data": []}
    assert unsupported == {"supported": False, "data": []}


def test_top_latency_volume_filter_excludes_non_storage_space_paths(db_session):
    db_session.add(
        models.StorageCluster(
            id=7,
            name="isilon-a",
            storage_type="isilon",
            storage_host="storage.local",
            is_active=True,
        )
    )
    db_session.add(
        models.Volume(
            storage_cluster_id=7,
            name="/ifs/data/project-a",
            performance_object_id="quota-42",
            vserver="",
            aggregate="",
            type="directory_quota",
            state="",
            updated_at=START,
        )
    )
    db_session.commit()
    analytics = _analytics()
    with patch.object(
        analytics.storageHealthAnalyticsCrud,
        "get_top_latency_rows",
        return_value=[
            {
                "object_id": object_id,
                "object_name": object_name,
                "object_type": "volume",
                "p95_latency": latency,
            }
            for object_id, object_name, latency in (
                ("unmatched-object", "/ifs/data/parent", 20.0),
                ("quota-42", "/ifs/data/project-a", 10.0),
            )
        ],
    ):
        result = analytics.get_top_latency(
            db_session,
            7,
            START,
            END,
            object_type="volume",
        )

    assert [row["object_name"] for row in result["data"]] == [
        "/ifs/data/project-a"
    ]


@pytest.mark.parametrize(
    ("start_time", "end_time", "message"),
    [
        (START, START, "start_time must be before end_time"),
        (START, START + timedelta(days=181), "180 days"),
    ],
)
def test_analytics_time_range_validation(start_time, end_time, message):
    with pytest.raises(ValueError, match=message):
        _analytics().validate_time_range(start_time, end_time)


def test_analytics_time_range_rejects_mixed_timezone_awareness_as_value_error():
    with pytest.raises(ValueError, match="timezone"):
        _analytics().validate_time_range(START, END.replace(tzinfo=None))


@pytest.fixture
def analytics_client(api_client_factory, auth_headers, db_session):
    db_session.add_all(
        [
            models.User(id=1, rd_username="alice"),
            models.StorageCluster(
                id=1,
                name="cluster-a",
                storage_type="netapp",
                storage_host="storage.local",
                is_active=True,
            ),
        ]
    )
    db_session.commit()
    return api_client_factory([storage_cluster.router], headers=auth_headers)


@pytest.mark.parametrize(
    ("endpoint", "service_name", "payload"),
    [
        ("capacity-change", "get_capacity_change", {"start_used": None, "points": []}),
        ("error-severity", "get_error_severity", {"counts": {}, "total": 0}),
        ("top-latency", "get_top_latency", []),
        ("repeated-faults", "get_repeated_faults", []),
        ("system-events", "get_system_events", {"data": []}),
    ],
)
def test_storage_health_read_endpoints_return_service_results(
    analytics_client, endpoint, service_name, payload
):
    with patch(f"routers.storage_cluster.{service_name}", return_value=payload):
        response = analytics_client.get(
            f"/storage-pulse/api/storage-clusters/1/analytics/{endpoint}",
            params=_query_params(),
        )

    assert response.status_code == 200
    body = response.json()
    if endpoint == "capacity-change":
        assert body.pop("trend_meta") == {
            "quota_basis": "hard",
            "rule_source": "system",
            "thresholds": {"important": 80, "serious": 90, "emergency": 95},
            "quota_limit_gb": None,
            "quota_limit_tb": None,
            "ratio_indicator": "used_ratio",
            "capacity": {},
        }
    assert body == payload


def test_storage_health_endpoint_defaults_to_the_previous_24_hours(analytics_client, monkeypatch):
    default_end = datetime(2026, 7, 23, 9, 30, tzinfo=timezone.utc)
    monkeypatch.setattr(storage_cluster, "_analytics_default_end", lambda: default_end)
    with patch("routers.storage_cluster.get_error_severity", return_value={"counts": {}, "total": 0}) as service:
        response = analytics_client.get(
            "/storage-pulse/api/storage-clusters/1/analytics/error-severity",
        )

    assert response.status_code == 200
    assert service.call_args.args[2:] == (default_end - timedelta(hours=24), default_end)


def test_system_events_endpoint_forwards_search_and_pagination(analytics_client):
    payload = {"data": [], "total": 0, "page": 2, "page_size": 20}
    with patch(
        "routers.storage_cluster.get_system_events", return_value=payload
    ) as service:
        response = analytics_client.get(
            "/storage-pulse/api/storage-clusters/1/analytics/system-events",
            params=_query_params(
                keyword="quota",
                severity="warning",
                page=2,
                page_size=20,
            ),
        )

    assert response.status_code == 200
    assert response.json() == payload
    service.assert_called_once_with(
        ANY,
        1,
        START,
        END,
        keyword="quota",
        severity="warning",
        page=2,
        page_size=20,
    )


def test_system_events_endpoint_rejects_unknown_severity(analytics_client):
    response = analytics_client.get(
        "/storage-pulse/api/storage-clusters/1/analytics/system-events",
        params=_query_params(severity="fatal"),
    )

    assert response.status_code == 422


@pytest.mark.parametrize(
    "params",
    [
        _query_params(end_time=START.isoformat()),
        _query_params(end_time=(START + timedelta(days=181)).isoformat()),
    ],
    ids=["reversed", "over-180-days"],
)
def test_storage_health_endpoint_rejects_invalid_time_ranges(analytics_client, params):
    response = analytics_client.get(
        "/storage-pulse/api/storage-clusters/1/analytics/capacity-change",
        params=params,
    )

    assert response.status_code == 422


def test_storage_health_endpoint_returns_422_for_mixed_timezone_awareness(analytics_client):
    response = analytics_client.get(
        "/storage-pulse/api/storage-clusters/1/analytics/capacity-change",
        params={
            "start_time": START.isoformat(),
            "end_time": END.replace(tzinfo=None).isoformat(),
        },
    )

    assert response.status_code == 422


def test_top_latency_endpoint_supports_up_to_one_hundred_rows(analytics_client):
    with patch("routers.storage_cluster.get_top_latency", return_value={"data": []}) as service:
        supported = analytics_client.get(
            "/storage-pulse/api/storage-clusters/1/analytics/top-latency",
            params=_query_params(limit=100),
        )
        rejected = analytics_client.get(
            "/storage-pulse/api/storage-clusters/1/analytics/top-latency",
            params=_query_params(limit=101),
        )

    assert supported.status_code == 200
    assert rejected.status_code == 422
    assert service.call_args.kwargs["limit"] == 100


@pytest.mark.parametrize(
    ("export_format", "section", "media_type", "filename"),
    [
        ("csv", "capacity", "text/csv", "storage-health-capacity.csv"),
        (
            "excel",
            "all",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "storage-health.xlsx",
        ),
        ("pdf", "all", "application/pdf", "storage-health.pdf"),
        ("csv", "all", "application/zip", "storage-health-csv.zip"),
    ],
)
def test_storage_health_export_media_contract(
    analytics_client, export_format, section, media_type, filename
):
    with patch(
        "routers.storage_cluster.export_storage_health",
        return_value=(b"report", media_type, filename),
    ):
        response = analytics_client.get(
            "/storage-pulse/api/storage-clusters/1/analytics/export",
            params=_query_params(format=export_format, section=section),
        )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(media_type)
    assert filename in response.headers["content-disposition"]


def test_pdf_export_succeeds_when_backend_logo_is_missing(monkeypatch, tmp_path):
    analytics = _analytics()
    missing_logo = tmp_path / "missing-logo.png"

    class OptionalLogoPDF:
        def __init__(self, *, logo_path, **_kwargs):
            if logo_path is not None and not Path(logo_path).exists():
                raise FileNotFoundError(logo_path)

        def create_cover_page(self):
            return None

        def add_table(self, *_args):
            return None

        def generate_pdf(self):
            return io.BytesIO(b"%PDF-1.4\n")

    monkeypatch.setattr(Config, "app_logo_path", property(lambda _self: missing_logo))
    monkeypatch.setattr(analytics, "PDFReportGenerator", OptionalLogoPDF)

    content = analytics._pdf_bytes(
        ["capacity"],
        {
            "capacity": {
                "start_used": None,
                "end_used": None,
                "change": None,
                "change_percent": None,
                "data": [],
            }
        },
    )

    assert content.startswith(b"%PDF")


def test_capacity_and_severity_section_rows_include_export_summaries():
    analytics = _analytics()
    sources = {
        "netapp": {"critical": 1, "error": 0, "warning": 0, "info": 0}
    }
    report = {
        "capacity": {
            "start_used": 100.0,
            "end_used": 125.0,
            "change": 25.0,
            "change_percent": 25.0,
            "data": [{"updated_at": START, "used": 100.0}],
        },
        "severity": {
            "counts": {"critical": 1, "error": 0, "warning": 0, "info": 0},
            "total": 1,
            "sources": sources,
        },
    }

    capacity_rows = analytics._section_rows("capacity", report)
    severity_rows = analytics._section_rows("severity", report)

    assert {
        "start_used": 100.0,
        "end_used": 125.0,
        "change": 25.0,
        "change_percent": 25.0,
    }.items() <= capacity_rows[0].items()
    assert {"updated_at": START, "used": 100.0}.items() <= capacity_rows[0].items()
    assert severity_rows[0]["total"] == 1
    assert severity_rows[0]["sources"] == sources


@pytest.mark.parametrize("prefix", ["=", "+", "-", "@"])
def test_csv_export_escapes_formula_prefixes(prefix):
    untrusted = f"{prefix}malicious"

    rows = list(
        csv.DictReader(
            io.StringIO(
                _analytics()._csv_bytes([{"object_name": untrusted}]).decode("utf-8-sig")
            )
        )
    )

    assert rows[0]["object_name"] == f"'{untrusted}"


@pytest.mark.parametrize("prefix", ["=", "+", "-", "@"])
def test_excel_export_escapes_formula_prefixes(prefix):
    untrusted = f"{prefix}malicious"
    content = _analytics()._excel_bytes(
        ["faults"],
        {"faults": {"data": [{"object_name": untrusted}]}},
    )

    workbook = load_workbook(io.BytesIO(content), data_only=False)

    assert workbook["faults"]["A2"].value == f"'{untrusted}"


@pytest.mark.parametrize(
    "untrusted",
    ["  =malicious", "\t+malicious", "\r-malicious", "\n@malicious"],
)
def test_csv_export_escapes_formula_after_leading_whitespace(untrusted):
    rows = list(
        csv.DictReader(
            io.StringIO(
                _analytics()._csv_bytes([{"object_name": untrusted}]).decode("utf-8-sig")
            )
        )
    )

    assert rows[0]["object_name"] == f"'{untrusted}"


@pytest.mark.parametrize(
    "untrusted",
    ["  =malicious", "\t+malicious", "\r-malicious", "\n@malicious"],
)
def test_excel_export_escapes_formula_after_leading_whitespace(untrusted):
    content = _analytics()._excel_bytes(
        ["faults"],
        {"faults": {"data": [{"object_name": untrusted}]}},
    )

    workbook = load_workbook(io.BytesIO(content), data_only=False)

    expected = ("'" + untrusted).replace("\r", "\n")
    assert workbook["faults"]["A2"].value == expected


def test_netapp_client_exposes_ems_and_volume_metrics_contracts():
    client = object.__new__(NetAppClient)
    client._get_all_records = Mock(side_effect=[[{"index": 7}], [{"uuid": "volume-1"}]])
    since = "2026-07-15T00:00:00Z"

    assert client.get_ems_events(since) == [{"index": 7}]
    assert client.get_volume_metrics() == [{"uuid": "volume-1"}]

    event_call, metrics_call = client._get_all_records.call_args_list
    assert event_call.args[0] == "support/ems/events"
    assert event_call.kwargs["params"]["time"] == f">={since}"
    assert {"time", "index", "message", "node", "log_message"} <= set(
        event_call.kwargs["params"]["fields"].split(",")
    )
    assert metrics_call.args[0] == "storage/volumes"
    assert {"uuid", "name", "metric"} <= set(metrics_call.kwargs["params"]["fields"].split(","))
    assert "metrics" not in metrics_call.kwargs["params"]["fields"].split(",")


def test_isilon_client_discovers_platform_version_before_statistics_and_events():
    client = object.__new__(IsilonClient)
    client.api_version = "1"
    client._get = Mock(
        side_effect=[
            {"latest": 16},
            {
                "datasets": [
                    {
                        "id": 3,
                        "metrics": ["path"],
                        "statkey": "cluster.performance.dataset.3",
                    }
                ]
            },
            {"workloads": []},
            {"stats": []},
            {"eventgroups": [{"id": "event-1"}]},
            {"eventlists": [{"id": "event-list-1"}]},
        ]
    )

    assert client.discover_api_version() == "16"
    assert client.get_performance_statistics() == []
    assert client.get_event_group_occurrences() == [{"id": "event-1"}]
    assert client.get_event_lists() == [{"id": "event-list-1"}]
    assert [call.args[0] for call in client._get.call_args_list] == [
        "/latest",
        "/16/performance/datasets",
        "/16/performance/datasets/3/workloads",
        "/16/statistics/current",
        "/16/event/eventgroup-occurrences",
        "/16/event/eventlists",
    ]
    assert client._get.call_args_list[3].kwargs["params"]["keys"] == "cluster.performance.dataset.3"


def test_isilon_statistics_requires_path_performance_dataset():
    client = object.__new__(IsilonClient)
    client.api_version = "16"
    client._get = Mock(return_value={"datasets": [{"id": 2, "metrics": ["username"]}]})

    with pytest.raises(ValueError, match="path performance dataset"):
        client.get_performance_statistics()
    assert [call.args[0] for call in client._get.call_args_list] == [
        "/16/performance/datasets",
    ]


def test_isilon_statistics_collects_path_dataset_workload_latency():
    client = object.__new__(IsilonClient)
    client.api_version = "22"
    client._get = Mock(
        side_effect=[
            {
                "datasets": [
                    {
                        "id": 3,
                        "name": "path",
                        "metrics": ["path"],
                        "statkey": "cluster.performance.dataset.3",
                    }
                ]
            },
            {
                "workloads": [
                    {
                        "id": 100,
                        "metric_values": {"path": "/ifs/data/project-a"},
                    }
                ]
            },
            {
                "stats": [
                    {
                        "key": "cluster.performance.dataset.3",
                        "time": 1784172559,
                        "value": {
                            "dataset": {
                                "workloads": {
                                    "workloads": [
                                        {
                                            "id": 100,
                                            "type": "pinned",
                                            "record": {
                                                "latency_read": {"sum": 3000, "count": 2},
                                                "latency_write": {"sum": 9000, "count": 3},
                                                "latency_other": {"sum": 0, "count": 0},
                                                "ops": {"sum": 50},
                                                "bytes_in": {"sum": 1024},
                                                "bytes_out": {"sum": 2048},
                                            },
                                        }
                                    ]
                                }
                            }
                        },
                    }
                ]
            },
        ]
    )

    assert client.get_performance_statistics() == [
        {
            "key": "cluster.performance.dataset.3.latency",
            "workload": "/ifs/data/project-a",
            "name": "/ifs/data/project-a",
            "value": 2400.0,
            "latency_read": 1500.0,
            "latency_write": 3000.0,
            "iops_total": 50.0,
            "throughput_total": 3072.0,
            "unit": "microseconds",
            "time": 1784172559,
        }
    ]
    assert [call.args[0] for call in client._get.call_args_list] == [
        "/22/performance/datasets",
        "/22/performance/datasets/3/workloads",
        "/22/statistics/current",
    ]
    assert client._get.call_args_list[2].kwargs["params"] == {
        "keys": "cluster.performance.dataset.3"
    }


def test_isilon_latency_rows_use_returned_workload_dimension_not_key_guessing():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")

    rows = storage_health._isilon_performance_rows(
        7,
        [
            {
                "key": "ifs.ops.latency",
                "workload": "workload-1",
                "value": 3.5,
                "unit": "milliseconds",
                "timestamp": "2026-07-15T10:00:00Z",
            },
            {
                "key": "ifs.node.latency",
                "devid": 2,
                "value": 2.5,
                "unit": "milliseconds",
                "timestamp": "2026-07-15T10:00:00Z",
            },
        ],
        datetime(2026, 7, 15, 10, 0, 0),
    )

    assert [(row["object_type"], row["object_id"]) for row in rows] == [
        ("volume", "workload-1")
    ]


@pytest.mark.parametrize(
    ("unit", "expected"),
    [
        ("microseconds", 2.5),
        ("us", 2.5),
        ("usec", 2.5),
        ("milliseconds", 2500.0),
        ("ms", 2500.0),
        ("seconds", 2500000.0),
    ],
)
def test_isilon_latency_rows_convert_only_known_units(unit, expected):
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")

    rows = storage_health._isilon_performance_rows(
        7,
        [
            {
                "key": "ifs.workload.latency",
                "workload": "workload-1",
                "value": 2500,
                "unit": unit,
                "timestamp": "2026-07-15T10:00:00Z",
            }
        ],
        datetime(2026, 7, 15, 10, 0),
    )

    assert rows[0]["latency_total"] == expected
    assert rows[0]["object_type"] == "volume"


def test_isilon_latency_rows_include_path_read_and_write_latency():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")

    rows = storage_health._isilon_performance_rows(
        7,
        [
            {
                "key": "cluster.performance.dataset.3.latency",
                "workload": "/ifs/data/project-a",
                "name": "/ifs/data/project-a",
                "value": 2400,
                "latency_read": 1500,
                "latency_write": 3000,
                "iops_total": 50,
                "throughput_total": 3072,
                "unit": "microseconds",
                "time": 1784172559,
            }
        ],
        datetime(2026, 7, 16, 10, 30),
    )

    assert rows == [
        {
            "storage_cluster_id": "7",
            "vendor": "isilon",
            "object_type": "volume",
            "object_id": "/ifs/data/project-a",
            "object_name": "/ifs/data/project-a",
            "latency_read": 1.5,
            "latency_write": 3.0,
            "latency_total": 2.4,
            "iops_total": 50.0,
            "throughput_total": 3072.0,
            "collected_at": datetime.fromtimestamp(
                1784172559, timezone.utc
            ).replace(tzinfo=None),
        }
    ]


def test_isilon_latency_rows_keep_only_directory_quota_paths():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")
    records = [
        {
            "key": "cluster.performance.dataset.3.latency",
            "workload": path,
            "value": 1000,
            "unit": "microseconds",
        }
        for path in ("/ifs/data/project-a", "/ifs/data/parent")
    ]

    rows = storage_health._isilon_performance_rows(
        7,
        records,
        datetime(2026, 7, 16, 10, 30),
        volume_paths={"/ifs/data/project-a"},
    )

    assert [row["object_name"] for row in rows] == ["/ifs/data/project-a"]


def test_isilon_latency_rows_keep_zero_seconds_and_use_device_timestamp():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")
    captured_at = 1784168941

    rows = storage_health._isilon_performance_rows(
        7,
        [
            {
                "key": "node.disk.access.latency.0",
                "devid": 1,
                "value": 0.0,
                "unit": "seconds",
                "time": captured_at,
            }
        ],
        datetime(2026, 7, 16, 10, 30),
    )

    assert rows[0]["latency_total"] == 0.0
    assert rows[0]["collected_at"] == (
        datetime.fromtimestamp(captured_at, timezone.utc).replace(tzinfo=None)
    )


@pytest.mark.parametrize("unit", [None, "", "minutes"])
def test_isilon_latency_rows_skip_missing_or_unknown_units(unit):
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")

    rows = storage_health._isilon_performance_rows(
        7,
        [
            {
                "key": "ifs.node.latency",
                "devid": 2,
                "value": 2.5,
                "unit": unit,
                "timestamp": "2026-07-15T10:00:00Z",
            }
        ],
        datetime(2026, 7, 15, 10, 0),
    )

    assert rows == []


def test_storage_alert_model_and_postgres_migration_contract():
    table = models.StorageAlerts.__table__
    assert {"storage_cluster_id", "source", "external_event_id", "fingerprint", "severity"} <= set(
        table.columns.keys()
    )
    assert table.c.storage_cluster_id.foreign_keys
    assert any(
        {column.name for column in constraint.columns}
        == {"storage_cluster_id", "source", "external_event_id"}
        for constraint in table.constraints
    )
    assert {
        ("storage_cluster_id", "updated_at"),
        ("severity", "updated_at"),
        ("fingerprint", "updated_at"),
    } <= {tuple(column.name for column in index.columns) for index in table.indexes}

    migrations = list((BACKEND_ROOT / "migrate" / "versions").glob("*storage_health*.py"))
    assert len(migrations) == 1, "expected one PostgreSQL storage health migration"
    spec = importlib.util.spec_from_file_location(migrations[0].stem, migrations[0])
    migration = importlib.util.module_from_spec(spec)
    assert spec is not None and spec.loader is not None
    spec.loader.exec_module(migration)
    assert migration.down_revision == "000000000003"

    output = io.StringIO()
    migration.op = Operations(
        MigrationContext.configure(
            dialect_name="postgresql",
            opts={"as_sql": True, "output_buffer": output},
        )
    )
    migration.upgrade()
    sql = " ".join(output.getvalue().lower().split())
    for token in ("storage_cluster_id", "source", "external_event_id", "fingerprint", "severity"):
        assert token in sql
    assert "unique" in sql
    assert sql.count("create index") >= 3


def test_storage_health_migration_compiles_offline_for_mysql():
    migration_path = next(
        (BACKEND_ROOT / "migrate" / "versions").glob("*storage_health*.py")
    )
    spec = importlib.util.spec_from_file_location(migration_path.stem, migration_path)
    migration = importlib.util.module_from_spec(spec)
    assert spec is not None and spec.loader is not None
    spec.loader.exec_module(migration)

    output = io.StringIO()
    migration.op = Operations(
        MigrationContext.configure(
            dialect_name="mysql",
            opts={"as_sql": True, "output_buffer": output},
        )
    )

    migration.upgrade()

    assert "alter table storage_alerts" in output.getvalue().lower()


def test_storage_health_string_lengths_are_explicit_and_match_mysql_migration():
    migration_path = next(
        (BACKEND_ROOT / "migrate" / "versions").glob("*storage_health*.py")
    )
    spec = importlib.util.spec_from_file_location(migration_path.stem, migration_path)
    migration = importlib.util.module_from_spec(spec)
    assert spec is not None and spec.loader is not None
    spec.loader.exec_module(migration)

    output = io.StringIO()
    migration.op = Operations(
        MigrationContext.configure(
            dialect_name="mysql",
            opts={"as_sql": True, "output_buffer": output},
        )
    )
    migration.upgrade()
    sql = " ".join(output.getvalue().lower().split())

    for column_name in ("source", "external_event_id", "fingerprint", "severity"):
        length = models.StorageAlerts.__table__.c[column_name].type.length
        assert isinstance(length, int) and length > 0
        assert re.search(
            rf"add column [`\"]?{column_name}[`\"]? varchar\({length}\)",
            sql,
        )


def test_storage_health_migration_backfills_attributable_alerts_on_sqlite():
    migration_paths = sorted((BACKEND_ROOT / "migrate" / "versions").glob("*.py"))
    migrations = []
    for path in migration_paths:
        spec = importlib.util.spec_from_file_location(path.stem, path)
        migration = importlib.util.module_from_spec(spec)
        assert spec is not None and spec.loader is not None
        spec.loader.exec_module(migration)
        migrations.append(migration)

    migrations_by_revision = {migration.revision: migration for migration in migrations}
    target_migration = migrations_by_revision["000000000004"]
    prior_migrations = []
    revision = target_migration.down_revision
    while revision is not None:
        migration = migrations_by_revision[revision]
        prior_migrations.append(migration)
        revision = migration.down_revision

    with sa.create_engine("sqlite://").begin() as connection:
        for migration in reversed(prior_migrations):
            migration.op = Operations(MigrationContext.configure(connection))
            migration.upgrade()
        connection.execute(
            sa.text(
                "INSERT INTO storage_clusters (id, name, storage_type) "
                "VALUES (1, 'cluster-a', 'netapp')"
            )
        )
        connection.execute(
            sa.text(
                "INSERT INTO storage_usages (id, storage_cluster_id) VALUES (1, 1)"
            )
        )
        connection.execute(
            sa.text(
                "INSERT INTO storage_alerts "
                "(id, alert_level, related_id, related_type) "
                "VALUES (1, 'high', 1, 'StorageUsage')"
            )
        )

        target_migration.op = Operations(MigrationContext.configure(connection))
        target_migration.upgrade()

        row = connection.execute(
            sa.text(
                "SELECT storage_cluster_id, source, severity FROM storage_alerts WHERE id = 1"
            )
        ).one()
        assert tuple(row) == (1, "diskpulse", "critical")


def test_questdb_performance_model_and_migration_contract():
    table = QuestDBBase.metadata.tables["storage_performance_metrics"]
    assert {
        "storage_cluster_id",
        "vendor",
        "object_type",
        "object_id",
        "object_name",
        "latency_read",
        "latency_write",
        "latency_total",
        "iops_total",
        "throughput_total",
        "collected_at",
    } <= set(table.columns.keys())

    migrations = list((BACKEND_ROOT / "questdb" / "migrations").glob("*_storage_performance_metrics.sql"))
    assert len(migrations) == 1, "expected one QuestDB performance migration"
    sql = " ".join(migrations[0].read_text(encoding="utf-8").lower().split())
    assert "create table" in sql
    assert "storage_performance_metrics" in sql
    assert "partition by day" in sql
    assert "wal" in sql
    assert "ttl 180 days" in sql


def test_celery_schedules_event_and_performance_collection():
    source = (BACKEND_ROOT / "celery_worker.py").read_text(encoding="utf-8")

    assert '"storage_events_schedule_fetching_task"' in source
    assert '"task": "celery_tasks.tasks.storage_health.storage_events_schedule_fetching_task"' in source
    assert '"schedule": 60.0' in source
    assert '"storage_performance_schedule_fetching_task"' in source
    assert '"task": "celery_tasks.tasks.storage_health.storage_performance_schedule_fetching_task"' in source
    assert '"schedule": 300.0' in source


def test_storage_event_window_uses_24_hours_then_five_minute_overlap():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")
    now = datetime(2026, 7, 15, 12, 0, 0)

    assert storage_health.event_window_start(None, now) == now - timedelta(hours=24)
    assert storage_health.event_window_start(
        datetime(2026, 7, 15, 11, 30, 0), now
    ) == datetime(2026, 7, 15, 11, 25, 0)


def test_netapp_since_requires_an_aware_utc_query_boundary():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")

    with pytest.raises(ValueError, match="timezone-aware"):
        storage_health._utc_z(datetime(2026, 7, 15, 10, 0))
    assert storage_health._utc_z(
        datetime(2026, 7, 15, 2, 0, tzinfo=timezone.utc)
    ) == "2026-07-15T02:00:00Z"


def test_netapp_first_event_collection_converts_local_now_to_utc_z(monkeypatch):
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")
    client = Mock()
    client.get_ems_events.return_value = []

    class LocalDatetime(datetime):
        @classmethod
        def now(cls, tz=None):
            value = datetime(
                2026,
                7,
                15,
                10,
                0,
                0,
                tzinfo=timezone(timedelta(hours=8)),
            )
            return value.astimezone(tz) if tz is not None else value

    class Result:
        def scalar_one_or_none(self):
            return None

    class Database:
        def execute(self, _statement):
            return Result()

        def commit(self):
            return None

        def rollback(self):
            return None

    class SessionContext:
        def __enter__(self):
            return Database()

        def __exit__(self, *_args):
            return False

    class Monitor:
        def __init__(self, *_args):
            self.storage_type = "netapp"
            self.client = client

        def setup(self):
            return None

        def cleanup(self):
            return None

    monkeypatch.setattr(storage_health, "SessionLocal", SessionContext)
    monkeypatch.setattr(storage_health, "StoragePulseMonitor", Monitor)
    monkeypatch.setattr(storage_health, "_persist_vendor_events", Mock(return_value=0))
    monkeypatch.setattr(
        storage_health,
        "utc_now",
        lambda: datetime(2026, 7, 15, 2, 0, 0, tzinfo=timezone.utc),
    )

    assert storage_health._collect_events(7) == 0
    client.get_ems_events.assert_called_once_with("2026-07-14T02:00:00Z")


def test_vendor_event_offset_timestamp_is_normalized_to_aware_utc():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")
    rows = storage_health.normalize_vendor_events(
        7,
        "netapp",
        [
            {
                "index": 42,
                "time": "2026-07-15T10:00:00+08:00",
                "message": {"name": "disk.offline", "severity": "error"},
                "node": {"uuid": "node-1"},
            }
        ],
    )

    assert rows[0]["updated_at"] == datetime(2026, 7, 15, 2, 0, 0, tzinfo=timezone.utc)


def test_vendor_event_parser_normalizes_identity_severity_and_fingerprint():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")
    rows = storage_health.normalize_vendor_events(
        storage_cluster_id=7,
        vendor="netapp",
        records=[
            {
                "index": 42,
                "time": "2026-07-15T10:00:00Z",
                "message": {"name": "disk.offline", "severity": "EMERGENCY"},
                "node": {"uuid": "node-1", "name": "node-a"},
                "log_message": "Disk offline",
            }
        ],
    )

    assert len(rows) == 1
    assert rows[0]["external_event_id"] == "node-1:42"
    assert rows[0]["severity"] == "critical"
    assert rows[0]["fingerprint"] == "netapp:disk.offline:node:node-1"
    assert rows[0]["storage_cluster_id"] == 7


def test_isilon_event_group_uses_last_event_and_cause_details():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")
    last_event = 1781582403

    rows = storage_health.normalize_vendor_events(
        7,
        "isilon",
        [
            {
                "id": "724916",
                "eventgroup_instance": "724916",
                "severity": "information",
                "time_noticed": 1780804835,
                "last_event": last_event,
                "causes": [
                    [
                        "QUOTA_THRESHOLD_VIOLATION",
                        "SmartQuotas threshold violation",
                    ]
                ],
                "specifier": {"devid": 6},
            }
        ],
    )

    assert rows[0]["external_event_id"] == "724916"
    assert rows[0]["fingerprint"] == "isilon:QUOTA_THRESHOLD_VIOLATION:node:6"
    assert rows[0]["description"] == "SmartQuotas threshold violation"
    assert rows[0]["updated_at"] == datetime.fromtimestamp(last_event, timezone.utc)


def test_isilon_event_list_normalizes_nested_events():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")
    event_time = 1783274393

    rows = storage_health.normalize_vendor_events(
        7,
        "isilon",
        [
            {
                "id": "737855",
                "events": [
                    {
                        "devid": 6,
                        "event": 600010001,
                        "id": "6.346657",
                        "message": "Snapshot creation failed",
                        "severity": "warning",
                        "time": event_time,
                    }
                ],
            }
        ],
    )

    assert len(rows) == 1
    assert rows[0]["external_event_id"] == "6.346657"
    assert rows[0]["fingerprint"] == "isilon:600010001:node:6"
    assert rows[0]["description"] == "Snapshot creation failed"
    assert rows[0]["updated_at"] == datetime.fromtimestamp(event_time, timezone.utc)


def test_storage_health_parser_boundary_values():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")
    now = datetime(2026, 7, 16, 10, 30, tzinfo=timezone.utc)

    assert storage_health._datetime(now) == now
    assert storage_health._datetime(None, now) == now
    assert storage_health._number("invalid") is None
    assert storage_health._netapp_performance_rows(7, [{"metric": {}}], now) == []
    assert storage_health._isilon_performance_rows(7, [{"key": "ifs.ops"}], now) == []
    assert storage_health._write_performance_rows([]) == 0
    assert storage_health._persist_vendor_events(Mock(), [], now) == 0
    assert len(storage_health._event_identity("isilon", {"severity": "info"})[0]) == 64


@pytest.mark.parametrize(
    ("vendor", "record"),
    [
        (
            "netapp",
            {
                "time": "2026-07-15T10:00:00Z",
                "message": {"name": "disk.offline", "severity": "error"},
            },
        ),
        (
            "netapp",
            {
                "index": 42,
                "message": {"name": "disk.offline", "severity": "error"},
            },
        ),
        (
            "netapp",
            {
                "index": 42,
                "time": "2026-07-15T10:00:00Z",
                "message": {"name": "disk.offline"},
            },
        ),
        (
            "isilon",
            {
                "start_time": "2026-07-15T10:00:00Z",
                "severity": "error",
                "event_type": "node.down",
            },
        ),
        (
            "isilon",
            {"id": "event-1", "severity": "error", "event_type": "node.down"},
        ),
        (
            "isilon",
            {
                "id": "event-1",
                "start_time": "2026-07-15T10:00:00Z",
                "event_type": "node.down",
            },
        ),
    ],
    ids=[
        "netapp-missing-id",
        "netapp-missing-time",
        "netapp-missing-severity",
        "isilon-missing-id",
        "isilon-missing-time",
        "isilon-missing-severity",
    ],
)
def test_vendor_events_missing_identity_time_or_severity_are_skipped(vendor, record):
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")

    assert storage_health.normalize_vendor_events(7, vendor, [record]) == []


def test_cluster_collection_isolates_failures():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")

    def collect(cluster_id):
        if cluster_id == 2:
            raise RuntimeError("offline")
        return cluster_id

    assert storage_health.run_isolated([1, 2, 3], collect, Mock()) == {
        "succeeded_clusters": (1, 3),
        "failed_clusters": (2,),
    }


def test_vendor_event_batch_deduplicates_external_event_id_before_insert(db_session):
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")
    event = {
        "index": 42,
        "time": "2026-07-15T10:00:00Z",
        "message": {"name": "disk.offline", "severity": "error"},
        "node": {"uuid": "node-1", "name": "node-a"},
        "log_message": "Disk offline",
    }
    rows = storage_health.normalize_vendor_events(7, "netapp", [event, event])

    inserted = storage_health._persist_vendor_events(
        db_session,
        rows,
        datetime(2026, 7, 15, 9, 55, 0, tzinfo=timezone.utc),
    )
    db_session.commit()

    assert inserted == 1
    assert db_session.query(models.StorageAlerts).count() == 1


def test_netapp_latency_metrics_are_converted_from_microseconds_to_milliseconds():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")

    rows = storage_health._netapp_performance_rows(
        7,
        [
            {
                "uuid": "volume-1",
                "name": "vol-a",
                "metric": {
                    "latency": {"total": 2500, "read": 1500, "write": 3500},
                    "iops": {"total": 125, "read": 75, "write": 50},
                    "throughput": {"total": 4096, "read": 3072, "write": 1024},
                    "timestamp": "2026-07-15T10:00:00Z",
                },
            }
        ],
        datetime(2026, 7, 15, 10, 0, 0),
    )

    assert rows[0]["latency_total"] == 2.5
    assert rows[0]["latency_read"] == 1.5
    assert rows[0]["latency_write"] == 3.5
    assert rows[0]["iops_total"] == 125.0
    assert rows[0]["throughput_total"] == 4096.0


def test_netapp_performance_collection_links_only_known_volume_identities():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")

    rows = storage_health._netapp_performance_rows(
        7,
        [
            {
                "uuid": "volume-1",
                "name": "vol-a",
                "metric": {"latency": {"total": 2500}},
            },
            {
                "uuid": "unknown-volume",
                "name": "vol-unknown",
                "metric": {"latency": {"total": 1800}},
            },
        ],
        datetime(2026, 7, 15, 10, 0, 0),
        volume_identities={"volume-1": "volume-1"},
    )

    assert [(row["object_id"], row["object_name"]) for row in rows] == [
        ("volume-1", "vol-a")
    ]


def test_isilon_performance_collection_uses_linked_volume_identity_and_drops_unmatched_workloads():
    storage_health = importlib.import_module("celery_tasks.tasks.storage_health")

    rows = storage_health._isilon_performance_rows(
        7,
        [
            {
                "key": "cluster.performance.dataset.3.latency",
                "workload": "/ifs/data/project-a",
                "value": 2400,
                "unit": "microseconds",
            },
            {
                "key": "cluster.performance.dataset.3.latency",
                "workload": "/ifs/data/unmatched",
                "value": 1800,
                "unit": "microseconds",
            },
        ],
        datetime(2026, 7, 16, 10, 30),
        volume_identities={"/ifs/data/project-a": "quota-42"},
    )

    assert [(row["object_id"], row["object_name"]) for row in rows] == [
        ("quota-42", "/ifs/data/project-a")
    ]
